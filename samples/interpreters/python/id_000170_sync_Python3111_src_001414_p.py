


"Software"), to deal








"AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR











import os.path
from xml.sax.saxutils import XMLGenerator


from openpyxl import Workbook
from openpyxl.compat import StringIO
from openpyxl.formatting import ConditionalFormatting
from openpyxl.formatting.rules import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.compat import iterkeys


from openpyxl.reader.excel import load_workbook
from openpyxl.reader.style import read_style_table
from openpyxl.xml.constants import ARC_STYLE
from openpyxl.writer.worksheet import write_worksheet_conditional_formatting
from openpyxl.writer.styles import StyleWriter
from openpyxl.styles import Border, Color, Fill, Font, Borders, HashableObject


import pytest
from zipfile import ZIP_DEFLATED, ZipFile
from openpyxl.tests.helper import DATADIR, get_xml, compare_xml


class TestRule:

    def test_ctor(self, FormatRule):
        r = FormatRule()
        assert r == {}

    @pytest.mark.parametrize("key, value",
                             [('aboveAverage', 1),
                              ('bottom', 0),
                              ('dxfId', True),
                              ('equalAverage', False),
                              ('operator', ""),
                              ('percent', 0),
                              ('priority', 1),
                              ('rank', 4),
                              ('stdDev', 2),
                              ('stopIfTrue', False),
                              ('text', "Once upon a time"),
                             ])
    def test_setitem(self, FormatRule, key, value):
        r1 = FormatRule()
        r2 = FormatRule()
        r1[key] = value
        setattr(r2, key, value)
        assert r1 == r2

    def test_getitem(self, FormatRule):
        r = FormatRule()
        r.aboveAverage = 1
        assert r.aboveAverage == r['aboveAverage']

    def test_invalid_key(self, FormatRule):
        r = FormatRule()
        with pytest.raises(KeyError):
            r['randomkey'] = 1
        with pytest.raises(KeyError):
            r['randomkey']

    def test_update_from_dict(self, FormatRule):
        r = FormatRule()
        d = {'aboveAverage':1}
        r.update(d)

    def test_len(self, FormatRule):
        r = FormatRule()
        assert len(r) == 0
        r.aboveAverage = 1
        assert len(r) == 1

    def test_keys(self, FormatRule):
        r = FormatRule()
        assert r.keys() == []
        r['operator'] = True
        assert r.keys() == ['operator']

    def test_values(self, FormatRule):
        r = FormatRule()
        assert r.values() == []
        r['rank'] = 1
        assert r.values() == [1]

    def test_items(self, FormatRule):
        r = FormatRule()
        assert r.items() == []
        r['stopIfTrue'] = False
        assert r.items() == [('stopIfTrue', False)]


class TestConditionalFormatting(object):

    class WB():
        style_properties = None
        worksheets = []

    def setup(self):
        self.workbook = self.WB()

    def test_conditional_formatting_customRule(self):
        class WS():
            conditional_formatting = ConditionalFormatting()
        worksheet = WS()
        worksheet.conditional_formatting.add('C1:C10', {'type': 'expression', 'formula': ['ISBLANK(C1)'],
                                                        'stopIfTrue': '1', 'dxf': {}})
        worksheet.conditional_formatting.setDxfStyles(self.workbook)
        temp_buffer = StringIO()
        doc = XMLGenerator(out=temp_buffer, encoding='utf-8')
        write_worksheet_conditional_formatting(doc, worksheet)
        doc.endDocument()
        xml = temp_buffer.getvalue()
        temp_buffer.close()

        diff = compare_xml(xml, """
        <conditionalFormatting sqref="C1:C10">
          <cfRule dxfId="0" type="expression" stopIfTrue="1" priority="1">
            <formula>ISBLANK(C1)</formula>
          </cfRule>
        </conditionalFormatting>
        """)
        assert diff is None, diff

    def test_conditional_formatting_setDxfStyle(self):
        cf = ConditionalFormatting()
        fill = Fill()
        fill.start_color.index = 'FFEE1111'
        fill.end_color.index = 'FFEE1111'
        fill.fill_type = Fill.FILL_SOLID
        font = Font()
        font.name = 'Arial'
        font.size = 12
        font.bold = True
        font.underline = Font.UNDERLINE_SINGLE
        borders = Borders()
        borders.top.border_style = Border.BORDER_THIN
        borders.top.color.index = Color.DARKYELLOW
        borders.bottom.border_style = Border.BORDER_THIN
        borders.bottom.color.index = Color.BLACK
        cf.add('C1:C10', FormulaRule(formula=['ISBLANK(C1)'], font=font, border=borders, fill=fill))
        cf.add('D1:D10', FormulaRule(formula=['ISBLANK(D1)'], fill=fill))
        cf.setDxfStyles(self.workbook)
        assert len(self.workbook.style_properties['dxf_list']) == 2
        assert self.workbook.style_properties['dxf_list'][0] == {'font': font, 'border': borders, 'fill': fill}
        assert self.workbook.style_properties['dxf_list'][1] == {'fill': fill}

    def test_conditional_formatting_update(self):
        class WS():
            conditional_formatting = ConditionalFormatting()
        worksheet = WS()
        rules = {'A1:A4': [{'type': 'colorScale', 'priority': 13,
                            'colorScale': {'cfvo': [{'type': 'min'}, {'type': 'max'}], 'color':
                                           [Color('FFFF7128'), Color('FFFFEF9C')]}}]}
        worksheet.conditional_formatting.update(rules)

        temp_buffer = StringIO()
        doc = XMLGenerator(out=temp_buffer, encoding='utf-8')
        write_worksheet_conditional_formatting(doc, worksheet)
        doc.endDocument()
        xml = temp_buffer.getvalue()
        temp_buffer.close()
        diff = compare_xml(xml, """
        <conditionalFormatting sqref="A1:A4">
          <cfRule type="colorScale" priority="1">
            <colorScale>
              <cfvo type="min" />
              <cfvo type="max" />
              <color rgb="FFFF7128" />
              <color rgb="FFFFEF9C" />
            </colorScale>
          </cfRule>
        </conditionalFormatting>
        """)
        assert diff is None, diff

    def test_conditional_font(self):
        """Test to verify font style written correctly."""
        class WS():
            conditional_formatting = ConditionalFormatting()
        worksheet = WS()

        
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        whiteFont = Font()
        whiteFont.color.index = "FFFFFFFF"
        worksheet.conditional_formatting.add('A1:A3', CellIsRule(operator='equal', formula=['"Fail"'], stopIfTrue=False,
                                                                 font=whiteFont, fill=redFill))
        worksheet.conditional_formatting.setDxfStyles(self.workbook)

        
        temp_buffer = StringIO()
        doc = XMLGenerator(out=temp_buffer, encoding='utf-8')
        write_worksheet_conditional_formatting(doc, worksheet)
        doc.endDocument()
        xml = temp_buffer.getvalue()
        temp_buffer.close()
        diff = compare_xml(xml, """
        <conditionalFormatting sqref="A1:A3">
          <cfRule dxfId="0" operator="equal" priority="1" type="cellIs">
            <formula>"Fail"</formula>
          </cfRule>
        </conditionalFormatting>
        """)
        assert diff is None, diff

        
        w = StyleWriter(self.workbook)
        w._write_dxfs()
        xml = get_xml(w._root)
        diff = compare_xml(xml, """
        <styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
          <dxfs count="1">
            <dxf>
              <font>
                <color rgb="FFFFFFFF" />
              </font>
              <fill>
                <patternFill patternType="solid">
                  <fgColor rgb="FFEE1111" />
                  <bgColor rgb="FFEE1111" />
                </patternFill>
              </fill>
            </dxf>
          </dxfs>
        </styleSheet>
        """)
        assert diff is None, diff


class TestColorScaleRule(object):

    class WB():
        style_properties = None
        worksheets = []

    def setup(self):
        self.workbook = self.WB()

    def test_two_colors(self):
        cf = ConditionalFormatting()

        cfRule = ColorScaleRule(start_type='min', start_value=None, start_color='FFAA0000',
                                end_type='max', end_value=None, end_color='FF00AA00')
        cf.add('A1:A10', cfRule)
        rules = cf.cf_rules
        assert 'A1:A10' in rules
        assert len(cf.cf_rules['A1:A10']) == 1
        assert rules['A1:A10'][0]['priority'] == 1
        assert rules['A1:A10'][0]['type'] == 'colorScale'
        assert rules['A1:A10'][0]['colorScale']['cfvo'][0]['type'] == 'min'
        assert rules['A1:A10'][0]['colorScale']['cfvo'][1]['type'] == 'max'

    def test_three_colors(self):
        cf = ConditionalFormatting()
        cfRule = ColorScaleRule(start_type='percentile', start_value=10, start_color='FFAA0000',
                                mid_type='percentile', mid_value=50, mid_color='FF0000AA',
                                end_type='percentile', end_value=90, end_color='FF00AA00')
        cf.add('B1:B10', cfRule)
        rules = cf.cf_rules
        assert 'B1:B10' in rules
        assert len(cf.cf_rules['B1:B10']) == 1
        assert rules['B1:B10'][0]['priority'] == 1
        assert rules['B1:B10'][0]['type'] == 'colorScale'
        assert rules['B1:B10'][0]['colorScale']['cfvo'][0]['type'] == 'percentile'
        assert rules['B1:B10'][0]['colorScale']['cfvo'][0]['val'] == '10'
        assert rules['B1:B10'][0]['colorScale']['cfvo'][1]['type'] == 'percentile'
        assert rules['B1:B10'][0]['colorScale']['cfvo'][1]['val'] == '50'
        assert rules['B1:B10'][0]['colorScale']['cfvo'][2]['type'] == 'percentile'
        assert rules['B1:B10'][0]['colorScale']['cfvo'][2]['val'] == '90'


class TestCellIsRule(object):

    class WB():
        style_properties = None
        worksheets = []

    def setup(self):
        self.workbook = self.WB()

    def test_greaterThan(self):
        cf = ConditionalFormatting()
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        cf.add('U10:U18', CellIsRule(operator='greaterThan', formula=['U$7'], stopIfTrue=True, fill=redFill))
        cf.add('V10:V18', CellIsRule(operator='>', formula=['V$7'], stopIfTrue=True, fill=redFill))
        cf.setDxfStyles(self.workbook)
        rules = cf.cf_rules
        assert 'U10:U18' in rules
        assert len(cf.cf_rules['U10:U18']) == 1
        assert rules['U10:U18'][0]['priority'] == 1
        assert rules['U10:U18'][0]['type'] == 'cellIs'
        assert rules['U10:U18'][0]['dxfId'] == 0
        assert rules['U10:U18'][0]['operator'] == 'greaterThan'
        assert rules['U10:U18'][0]['formula'][0] == 'U$7'
        assert rules['U10:U18'][0]['stopIfTrue'] == '1'
        assert 'V10:V18' in rules
        assert len(cf.cf_rules['V10:V18']) == 1
        assert rules['V10:V18'][0]['priority'] == 2
        assert rules['V10:V18'][0]['type'] == 'cellIs'
        assert rules['V10:V18'][0]['dxfId'] == 1
        assert rules['V10:V18'][0]['operator'] == 'greaterThan'
        assert rules['V10:V18'][0]['formula'][0] == 'V$7'
        assert rules['V10:V18'][0]['stopIfTrue'] == '1'

    def test_greaterThanOrEqual(self):
        cf = ConditionalFormatting()
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        cf.add('U10:U18', CellIsRule(operator='greaterThanOrEqual', formula=['U$7'], stopIfTrue=True, fill=redFill))
        cf.add('V10:V18', CellIsRule(operator='>=', formula=['V$7'], stopIfTrue=True, fill=redFill))
        cf.setDxfStyles(self.workbook)
        rules = cf.cf_rules
        assert 'U10:U18' in rules
        assert len(cf.cf_rules['U10:U18']) == 1
        assert rules['U10:U18'][0]['priority'] == 1
        assert rules['U10:U18'][0]['type'] == 'cellIs'
        assert rules['U10:U18'][0]['dxfId'] == 0
        assert rules['U10:U18'][0]['operator'] == 'greaterThanOrEqual'
        assert rules['U10:U18'][0]['formula'][0] == 'U$7'
        assert rules['U10:U18'][0]['stopIfTrue'] == '1'
        assert 'V10:V18' in rules
        assert len(cf.cf_rules['V10:V18']) == 1
        assert rules['V10:V18'][0]['priority'] == 2
        assert rules['V10:V18'][0]['type'] == 'cellIs'
        assert rules['V10:V18'][0]['dxfId'] == 1
        assert rules['V10:V18'][0]['operator'] == 'greaterThanOrEqual'
        assert rules['V10:V18'][0]['formula'][0] == 'V$7'
        assert rules['V10:V18'][0]['stopIfTrue'] == '1'

    def test_lessThan(self):
        cf = ConditionalFormatting()
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        cf.add('U10:U18', CellIsRule(operator='lessThan', formula=['U$7'], stopIfTrue=True, fill=redFill))
        cf.add('V10:V18', CellIsRule(operator='<', formula=['V$7'], stopIfTrue=True, fill=redFill))
        cf.setDxfStyles(self.workbook)
        rules = cf.cf_rules
        assert 'U10:U18' in rules
        assert len(cf.cf_rules['U10:U18']) == 1
        assert rules['U10:U18'][0]['priority'] == 1
        assert rules['U10:U18'][0]['type'] == 'cellIs'
        assert rules['U10:U18'][0]['dxfId'] == 0
        assert rules['U10:U18'][0]['operator'] == 'lessThan'
        assert rules['U10:U18'][0]['formula'][0] == 'U$7'
        assert rules['U10:U18'][0]['stopIfTrue'] == '1'
        assert 'V10:V18' in rules
        assert len(cf.cf_rules['V10:V18']) == 1
        assert rules['V10:V18'][0]['priority'] == 2
        assert rules['V10:V18'][0]['type'] == 'cellIs'
        assert rules['V10:V18'][0]['dxfId'] == 1
        assert rules['V10:V18'][0]['operator'] == 'lessThan'
        assert rules['V10:V18'][0]['formula'][0] == 'V$7'
        assert rules['V10:V18'][0]['stopIfTrue'] == '1'

    def test_lessThanOrEqual(self):
        cf = ConditionalFormatting()
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        cf.add('U10:U18', CellIsRule(operator='lessThanOrEqual', formula=['U$7'], stopIfTrue=True, fill=redFill))
        cf.add('V10:V18', CellIsRule(operator='<=', formula=['V$7'], stopIfTrue=True, fill=redFill))
        cf.setDxfStyles(self.workbook)
        rules = cf.cf_rules
        assert 'U10:U18' in rules
        assert len(cf.cf_rules['U10:U18']) == 1
        assert rules['U10:U18'][0]['priority'] == 1
        assert rules['U10:U18'][0]['type'] == 'cellIs'
        assert rules['U10:U18'][0]['dxfId'] == 0
        assert rules['U10:U18'][0]['operator'] == 'lessThanOrEqual'
        assert rules['U10:U18'][0]['formula'][0] == 'U$7'
        assert rules['U10:U18'][0]['stopIfTrue'] == '1'
        assert 'V10:V18' in rules
        assert len(cf.cf_rules['V10:V18']) == 1
        assert rules['V10:V18'][0]['priority'] == 2
        assert rules['V10:V18'][0]['type'] == 'cellIs'
        assert rules['V10:V18'][0]['dxfId'] == 1
        assert rules['V10:V18'][0]['operator'] == 'lessThanOrEqual'
        assert rules['V10:V18'][0]['formula'][0] == 'V$7'
        assert rules['V10:V18'][0]['stopIfTrue'] == '1'

    def test_equal(self):
        cf = ConditionalFormatting()
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        cf.add('U10:U18', CellIsRule(operator='equal', formula=['U$7'], stopIfTrue=True, fill=redFill))
        cf.add('V10:V18', CellIsRule(operator='=', formula=['V$7'], stopIfTrue=True, fill=redFill))
        cf.add('W10:W18', CellIsRule(operator='==', formula=['W$7'], stopIfTrue=True, fill=redFill))
        cf.setDxfStyles(self.workbook)
        rules = cf.cf_rules
        assert 'U10:U18' in rules
        assert len(cf.cf_rules['U10:U18']) == 1
        assert rules['U10:U18'][0]['priority'] == 1
        assert rules['U10:U18'][0]['type'] == 'cellIs'
        assert rules['U10:U18'][0]['dxfId'] == 0
        assert rules['U10:U18'][0]['operator'] == 'equal'
        assert rules['U10:U18'][0]['formula'][0] == 'U$7'
        assert rules['U10:U18'][0]['stopIfTrue'] == '1'
        assert 'V10:V18' in rules
        assert len(cf.cf_rules['V10:V18']) == 1
        assert rules['V10:V18'][0]['priority'] == 2
        assert rules['V10:V18'][0]['type'] == 'cellIs'
        assert rules['V10:V18'][0]['dxfId'] == 1
        assert rules['V10:V18'][0]['operator'] == 'equal'
        assert rules['V10:V18'][0]['formula'][0] == 'V$7'
        assert rules['V10:V18'][0]['stopIfTrue'] == '1'
        assert 'W10:W18' in rules
        assert len(cf.cf_rules['W10:W18']) == 1
        assert rules['W10:W18'][0]['priority'] == 3
        assert rules['W10:W18'][0]['type'] == 'cellIs'
        assert rules['W10:W18'][0]['dxfId'] == 2
        assert rules['W10:W18'][0]['operator'] == 'equal'
        assert rules['W10:W18'][0]['formula'][0] == 'W$7'
        assert rules['W10:W18'][0]['stopIfTrue'] == '1'

    def test_notEqual(self):
        cf = ConditionalFormatting()
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        cf.add('U10:U18', CellIsRule(operator='notEqual', formula=['U$7'], stopIfTrue=True, fill=redFill))
        cf.add('V10:V18', CellIsRule(operator='!=', formula=['V$7'], stopIfTrue=True, fill=redFill))
        cf.setDxfStyles(self.workbook)
        rules = cf.cf_rules
        assert 'U10:U18' in rules
        assert len(cf.cf_rules['U10:U18']) == 1
        assert rules['U10:U18'][0]['priority'] == 1
        assert rules['U10:U18'][0]['type'] == 'cellIs'
        assert rules['U10:U18'][0]['dxfId'] == 0
        assert rules['U10:U18'][0]['operator'] == 'notEqual'
        assert rules['U10:U18'][0]['formula'][0] == 'U$7'
        assert rules['U10:U18'][0]['stopIfTrue'] == '1'
        assert 'V10:V18' in rules
        assert len(cf.cf_rules['V10:V18']) == 1
        assert rules['V10:V18'][0]['priority'] == 2
        assert rules['V10:V18'][0]['type'] == 'cellIs'
        assert rules['V10:V18'][0]['dxfId'] == 1
        assert rules['V10:V18'][0]['operator'] == 'notEqual'
        assert rules['V10:V18'][0]['formula'][0] == 'V$7'
        assert rules['V10:V18'][0]['stopIfTrue'] == '1'

    def test_between(self):
        cf = ConditionalFormatting()
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        cf.add('U10:U18', CellIsRule(operator='between', formula=['U$7', 'U$8'], stopIfTrue=True, fill=redFill))
        cf.setDxfStyles(self.workbook)
        rules = cf.cf_rules
        assert 'U10:U18' in rules
        assert len(cf.cf_rules['U10:U18']) == 1
        assert rules['U10:U18'][0]['priority'] == 1
        assert rules['U10:U18'][0]['type'] == 'cellIs'
        assert rules['U10:U18'][0]['dxfId'] == 0
        assert rules['U10:U18'][0]['operator'] == 'between'
        assert rules['U10:U18'][0]['formula'][0] == 'U$7'
        assert rules['U10:U18'][0]['formula'][1] == 'U$8'
        assert rules['U10:U18'][0]['stopIfTrue'] == '1'

    def test_notBetween(self):
        cf = ConditionalFormatting()
        redFill = Fill()
        redFill.start_color.index = 'FFEE1111'
        redFill.end_color.index = 'FFEE1111'
        redFill.fill_type = Fill.FILL_SOLID
        cf.add('U10:U18', CellIsRule(operator='notBetween', formula=['U$7', 'U$8'], stopIfTrue=True, fill=redFill))
        cf.setDxfStyles(self.workbook)
        rules = cf.cf_rules
        assert 'U10:U18' in rules
        assert len(cf.cf_rules['U10:U18']) == 1
        assert rules['U10:U18'][0]['priority'] == 1
        assert rules['U10:U18'][0]['type'] == 'cellIs'
        assert rules['U10:U18'][0]['dxfId'] == 0
        assert rules['U10:U18'][0]['operator'] == 'notBetween'
        assert rules['U10:U18'][0]['formula'][0] == 'U$7'
        assert rules['U10:U18'][0]['formula'][1] == 'U$8'
        assert rules['U10:U18'][0]['stopIfTrue'] == '1'


class TestFormulaRule(object):
    class WB():
        style_properties = None
        worksheets = []

    def setup(self):
        self.workbook = self.WB()

    def test_formula_rule(self):
        class WS():
            conditional_formatting = ConditionalFormatting()
        worksheet = WS()
        worksheet.conditional_formatting.add('C1:C10', FormulaRule(formula=['ISBLANK(C1)'], stopIfTrue=True))
        worksheet.conditional_formatting.setDxfStyles(self.workbook)
        temp_buffer = StringIO()
        doc = XMLGenerator(out=temp_buffer, encoding='utf-8')
        write_worksheet_conditional_formatting(doc, worksheet)
        doc.endDocument()
        xml = temp_buffer.getvalue()
        temp_buffer.close()

        diff = compare_xml(xml, """
        <conditionalFormatting sqref="C1:C10">
          <cfRule dxfId="0" type="expression" stopIfTrue="1" priority="1">
            <formula>ISBLANK(C1)</formula>
          </cfRule>
        </conditionalFormatting>
        """)
        assert diff is None, diff

def compare_complex(a, b):
    if isinstance(a, list):
        if not isinstance(b, list):
            return False
        else:
            for i, v in enumerate(a):
                if not compare_complex(v, b[i]):
                    return False
    elif isinstance(a, dict):
        if not isinstance(b, dict):
            return False
        else:
            for k in iterkeys(a):
                if isinstance(a[k], (list, dict)):
                    if not compare_complex(a[k], b[k]):
                        return False
                elif a[k] != b[k]:
                    return False
    elif isinstance(a, HashableObject) or isinstance(b, HashableObject):
        if repr(a) != repr(b):
            return False
    elif a != b:
        return False
    return True


def test_conditional_formatting_read():
    reference_file = os.path.join(DATADIR, 'reader', 'conditional-formatting.xlsx')
    wb = load_workbook(reference_file)
    ws = wb.get_active_sheet()

    
    assert ws.conditional_formatting.cf_rules['A1:A1048576'] == [{'priority': 27, 'type': 'colorScale', 'colorScale': {'color': [Color('FFFF7128'), Color('FFFFEF9C')], 'cfvo': [{'type': 'min'}, {'type': 'max'}]}}]
    assert compare_complex(ws.conditional_formatting.cf_rules['B1:B10'], [{'priority': 26, 'type': 'colorScale', 'colorScale': {'color': [Color('theme:6:'), Color('theme:4:')], 'cfvo': [{'type': 'num', 'val': '3'}, {'type': 'num', 'val': '7'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['C1:C10'], [{'priority': 25, 'type': 'colorScale', 'colorScale': {'color': [Color('FFFF7128'), Color('FFFFEF9C')], 'cfvo': [{'type': 'percent', 'val': '10'}, {'type': 'percent', 'val': '90'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['D1:D10'], [{'priority': 24, 'type': 'colorScale', 'colorScale': {'color': [Color('theme:6:'), Color('theme:5:')], 'cfvo': [{'type': 'formula', 'val': '2'}, {'type': 'formula', 'val': '4'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['E1:E10'], [{'priority': 23, 'type': 'colorScale', 'colorScale': {'color': [Color('FFFF7128'), Color('FFFFEF9C')], 'cfvo': [{'type': 'percentile', 'val': '10'}, {'type': 'percentile', 'val': '90'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['F1:F10'], [{'priority': 22, 'type': 'colorScale', 'colorScale': {'color': [Color('FFFF7128'), Color('FFFFEB84'), Color('FF63BE7B')], 'cfvo': [{'type': 'min'}, {'type': 'percentile', 'val': '50'}, {'type': 'max'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['G1:G10'], [{'priority': 21, 'type': 'colorScale', 'colorScale': {'color': [Color('theme:4:'), Color('FFFFEB84'), Color('theme:5:')], 'cfvo': [{'type': 'num', 'val': '0'}, {'type': 'percentile', 'val': '50'}, {'type': 'num', 'val': '10'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['H1:H10'], [{'priority': 20, 'type': 'colorScale', 'colorScale': {'color': [Color('FFFF7128'), Color('FFFFEB84'), Color('FF63BE7B')], 'cfvo': [{'type': 'percent', 'val': '0'}, {'type': 'percent', 'val': '50'}, {'type': 'percent', 'val': '100'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['I1:I10'], [{'priority': 19, 'type': 'colorScale', 'colorScale': {'color': [Color('FF0000FF'), Color('FFFF6600'), Color('FF008000')], 'cfvo': [{'type': 'formula', 'val': '2'}, {'type': 'formula', 'val': '7'}, {'type': 'formula', 'val': '9'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['J1:J10'], [{'priority': 18, 'type': 'colorScale', 'colorScale': {'color': [Color('FFFF7128'), Color('FFFFEB84'), Color('FF63BE7B')], 'cfvo': [{'type': 'percentile', 'val': '10'}, {'type': 'percentile', 'val': '50'}, {'type': 'percentile', 'val': '90'}]}}])
    assert compare_complex(ws.conditional_formatting.cf_rules['K1:K10'], [])  
    assert compare_complex(ws.conditional_formatting.cf_rules['L1:L10'], [])  
    assert compare_complex(ws.conditional_formatting.cf_rules['M1:M10'], [])  
    assert compare_complex(ws.conditional_formatting.cf_rules['N1:N10'], [{'priority': 17, 'iconSet': {'cfvo': [{'type': 'percent', 'val': '0'}, {'type': 'percent', 'val': '33'}, {'type': 'percent', 'val': '67'}]}, 'type': 'iconSet'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['O1:O10'], [{'priority': 16, 'iconSet': {'cfvo': [{'type': 'percent', 'val': '0'}, {'type': 'num', 'val': '2'}, {'type': 'num', 'val': '4'}, {'type': 'num', 'val': '6'}], 'showValue': '0', 'iconSet': '4ArrowsGray', 'reverse': '1'}, 'type': 'iconSet'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['P1:P10'], [{'priority': 15, 'iconSet': {'cfvo': [{'type': 'percent', 'val': '0'}, {'type': 'percentile', 'val': '20'}, {'type': 'percentile', 'val': '40'}, {'type': 'percentile', 'val': '60'}, {'type': 'percentile', 'val': '80'}], 'iconSet': '5Rating'}, 'type': 'iconSet'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['Q1:Q10'], [{'text': '3', 'priority': 14, 'dxfId': '27', 'operator': 'containsText', 'formula': ['NOT(ISERROR(SEARCH("3",Q1)))'], 'type': 'containsText'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['R1:R10'], [{'operator': 'between', 'dxfId': '26', 'type': 'cellIs', 'formula': ['2', '7'], 'priority': 13}])
    assert compare_complex(ws.conditional_formatting.cf_rules['S1:S10'], [{'priority': 12, 'dxfId': '25', 'percent': '1', 'type': 'top10', 'rank': '10'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['T1:T10'], [{'priority': 11, 'dxfId': '24', 'type': 'top10', 'rank': '4', 'bottom': '1'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['U1:U10'], [{'priority': 10, 'dxfId': '23', 'type': 'aboveAverage'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['V1:V10'], [{'aboveAverage': '0', 'dxfId': '22', 'type': 'aboveAverage', 'priority': 9}])
    assert compare_complex(ws.conditional_formatting.cf_rules['W1:W10'], [{'priority': 8, 'dxfId': '21', 'type': 'aboveAverage', 'equalAverage': '1'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['X1:X10'], [{'aboveAverage': '0', 'dxfId': '20', 'priority': 7, 'type': 'aboveAverage', 'equalAverage': '1'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['Y1:Y10'], [{'priority': 6, 'dxfId': '19', 'type': 'aboveAverage', 'stdDev': '1'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['Z1:Z10'], [{'aboveAverage': '0', 'dxfId': '18', 'type': 'aboveAverage', 'stdDev': '1', 'priority': 5}])
    assert compare_complex(ws.conditional_formatting.cf_rules['AA1:AA10'], [{'priority': 4, 'dxfId': '17', 'type': 'aboveAverage', 'stdDev': '2'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['AB1:AB10'], [{'priority': 3, 'dxfId': '16', 'type': 'duplicateValues'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['AC1:AC10'], [{'priority': 2, 'dxfId': '15', 'type': 'uniqueValues'}])
    assert compare_complex(ws.conditional_formatting.cf_rules['AD1:AD10'], [{'priority': 1, 'dxfId': '14', 'type': 'expression', 'formula': ['AD1>3']}])


def test_parse_dxfs():
    reference_file = os.path.join(DATADIR, 'reader', 'conditional-formatting.xlsx')
    wb = load_workbook(reference_file)
    assert isinstance(wb, Workbook)
    archive = ZipFile(reference_file, 'r', ZIP_DEFLATED)
    read_xml = archive.read(ARC_STYLE)

    
    assert '<dxfs count="164">' in str(read_xml)
    assert len(wb.style_properties['dxf_list']) == 164

    
    reference_file = os.path.join(DATADIR, 'writer', 'expected', 'dxf_style.xml')
    with open(reference_file) as expected:
        diff = compare_xml(read_xml, expected.read())
        assert diff is None, diff

    cond_styles = wb.style_properties['dxf_list'][0]
    assert cond_styles['font'].color == Color('FF9C0006')
    assert not cond_styles['font'].bold
    assert not cond_styles['font'].italic
    f = Fill()
    f.end_color = Color('FFFFC7CE')
    assert cond_styles['fill'] == f

    
    w = StyleWriter(wb)
    w._write_dxfs()
    write_xml = get_xml(w._root)
    read_style_prop = read_style_table(write_xml)
    assert len(read_style_prop['dxf_list']) == len(wb.style_properties['dxf_list'])
    for i, dxf in enumerate(read_style_prop['dxf_list']):
        assert repr(wb.style_properties['dxf_list'][i] == dxf)
